import zipfile
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, List

import pandas as pd
import requests
from openpyxl import load_workbook

try:
    from etl.fetch.ans_indexer import get_last_trimesters
except ModuleNotFoundError:
    from ans_indexer import get_last_trimesters

DEFAULT_RAW_DIR = Path("data/raw/ans")
DEFAULT_EXTRACT_DIR = Path("data/processed/ans")
CHUNK_SIZE = 1024 * 1024
CSV_EXTS = {".csv", ".txt"}
XLSX_EXT = ".xlsx"
CSV_SEPS = [";", ",", "\t", "|", None]
CSV_ENCODINGS = ["utf-8", "latin-1"]


def _build_quarter_dir(ano: int, tri: int) -> str:
    """
    Monta o nome do diretorio no formato 3T2025.

    :param ano: Ano do trimestre.
    :param tri: Numero do trimestre.
    :return: Nome do diretorio.
    """
    return f"{tri}T{ano}"


def _zip_path(item: Dict, raw_dir: Path) -> Path:
    """
    Retorna o caminho de destino do ZIP para um item.

    :param item: Dados do trimestre.
    :param raw_dir: Diretorio base de download.
    :return: Caminho completo do ZIP.
    """
    ano = item["ano"]
    tri = item["trimestre"]
    return raw_dir / str(ano) / _build_quarter_dir(ano, tri) / item["filename"]


def download_file(url: str, dest: Path, timeout: int = 30) -> Path:
    """
    Baixa um arquivo para disco usando streaming.

    :param url: URL do arquivo.
    :param dest: Caminho de destino.
    :param timeout: Timeout em segundos.
    :return: Caminho do arquivo baixado.
    """
    dest.parent.mkdir(parents=True, exist_ok=True)
    with requests.get(url, stream=True, timeout=timeout) as r:
        r.raise_for_status()
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=CHUNK_SIZE):
                if chunk:
                    f.write(chunk)
    return dest


def download_items(items: List[Dict], raw_dir: Path = DEFAULT_RAW_DIR) -> List[Path]:
    """
    Baixa todos os ZIPs para os itens informados.

    :param items: Itens de trimestre.
    :param raw_dir: Diretorio base de download.
    :return: Lista de caminhos baixados.
    """
    paths: List[Path] = []
    for item in items:
        path = _zip_path(item, raw_dir)
        if path.exists():
            paths.append(path)
            continue
        paths.append(download_file(item["url"], path))
    return paths


def download_last_trimesters(
    limit: int = 3, raw_dir: Path = DEFAULT_RAW_DIR
) -> List[Path]:
    """
    Baixa os ZIPs dos ultimos N trimestres.

    :param limit: Quantidade de trimestres.
    :param raw_dir: Diretorio base de download.
    :return: Lista de caminhos baixados.
    """
    items = get_last_trimesters(limit)
    return download_items(items, raw_dir)


def extract_zip(zip_path: Path, dest_dir: Path) -> Path:
    """
    Extrai um ZIP para o diretorio de destino.

    :param zip_path: Caminho do ZIP.
    :param dest_dir: Diretorio de extracao.
    :return: Diretorio de extracao.
    """
    dest_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(dest_dir)
    return dest_dir


def extract_downloaded(
    items: List[Dict],
    raw_dir: Path = DEFAULT_RAW_DIR,
    extract_dir: Path = DEFAULT_EXTRACT_DIR,
) -> List[Path]:
    """
    Extrai todos os ZIPs baixados para os itens informados.

    :param items: Itens de trimestre.
    :param raw_dir: Diretorio base de download.
    :param extract_dir: Diretorio base de extracao.
    :return: Lista de diretorios extraidos.
    """
    extracted: List[Path] = []
    for item in items:
        ano = item["ano"]
        tri = item["trimestre"]
        zip_path = _zip_path(item, raw_dir)
        if not zip_path.exists():
            continue
        target_dir = (
            extract_dir / str(ano) / _build_quarter_dir(ano, tri) / zip_path.stem
        )
        extracted.append(extract_zip(zip_path, target_dir))
    return extracted


def _normalize_col_name(name: str) -> str:
    """
    Normaliza nome de coluna para comparar com DESCRICAO.

    :param name: Nome original da coluna.
    :return: Nome normalizado.
    """
    if name is None:
        return ""
    text = str(name).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _is_descricao_col(name: str) -> bool:
    """
    Retorna True se a coluna for DESCRICAO (normalizada).

    :param name: Nome da coluna.
    :return: True se for DESCRICAO.
    """
    return _normalize_col_name(name) == "descricao"


def _df_contains_evento(df: "pd.DataFrame") -> bool:
    """
    Verifica se DESCRICAO contem despesas de eventos/sinistros.

    :param df: DataFrame com dados.
    :return: True se encontrou eventos/sinistros.
    """
    if df is None or df.empty:
        return False
    descricao_cols = [col for col in df.columns if _is_descricao_col(col)]
    if not descricao_cols:
        return False
    series = df[descricao_cols[0]].astype(str).str.lower()
    mask = series.str.contains("despesa", na=False) & series.str.contains(
        r"(?:evento|sinistro)", na=False
    )
    return bool(mask.any())


def _csv_contains_evento(path: Path) -> bool:
    """
    Procura em CSV/TXT por despesas de eventos/sinistros.

    :param path: Caminho do arquivo.
    :return: True se encontrou eventos/sinistros.
    """
    for encoding in CSV_ENCODINGS:
        for sep in CSV_SEPS:
            try:
                reader = pd.read_csv(
                    path,
                    sep=sep,
                    engine="python",
                    dtype=str,
                    chunksize=50000,
                    encoding=encoding,
                    on_bad_lines="skip",
                    usecols=_is_descricao_col,
                )
                for chunk in reader:
                    if _df_contains_evento(chunk):
                        return True
                return False
            except Exception:
                continue
    return False


def _xlsx_contains_evento(path: Path) -> bool:
    """
    Procura em XLSX por despesas de eventos/sinistros.

    :param path: Caminho do arquivo.
    :return: True se encontrou eventos/sinistros.
    """
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            if not header:
                continue
            col_index = None
            for idx, name in enumerate(header):
                if _is_descricao_col(name):
                    col_index = idx
                    break
            if col_index is None:
                continue
            for row in rows:
                if not row or col_index >= len(row):
                    continue
                value = row[col_index]
                if value is None:
                    continue
                text = str(value).lower()
                if "despesa" in text and ("evento" in text or "sinistro" in text):
                    return True
        return False
    finally:
        wb.close()


def find_evento_files(paths: Iterable[Path]) -> List[Path]:
    """
    Retorna arquivos que contenham despesas de eventos/sinistros.

    :param paths: Diretorios base para busca.
    :return: Lista de arquivos encontrados.
    """
    results: List[Path] = []
    for root in paths:
        for file in root.rglob("*"):
            if not file.is_file():
                continue
            suffix = file.suffix.lower()
            if suffix in CSV_EXTS:
                if _csv_contains_evento(file):
                    results.append(file)
                continue
            if suffix == XLSX_EXT:
                if _xlsx_contains_evento(file):
                    results.append(file)
    return results


if __name__ == "__main__":
    downloaded = download_last_trimesters(3)
    for path in downloaded:
        print(path)
