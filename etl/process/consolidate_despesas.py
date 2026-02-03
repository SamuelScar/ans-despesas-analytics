import re
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


EXTRACT_DIR = Path("data/processed/ans")
CADOP_PATH = Path("Relatorio_cadop.csv")
OUTPUT_DIR = Path("data/output")
OUTPUT_FILE = OUTPUT_DIR / "consolidado_despesas.csv"

CSV_EXTS = {".csv", ".txt"}
XLSX_EXT = ".xlsx"
CSV_SEPS = [";", ",", "\t", "|", None]
CSV_ENCODINGS = ["utf-8", "latin-1"]
CADOP_ENCODINGS = ["utf-8-sig", "utf-8", "latin-1"]
QUARTER_RE = re.compile(r"([1-4])T(\d{4})", re.IGNORECASE)


def _normalize_key(value: object) -> str:
    """
    Normaliza texto para comparar colunas entre arquivos.

    :param value: Valor original.
    :return: Texto normalizado.
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(ch for ch in text if ch.isalnum())


def _find_col(
    columns: Iterable[str], keys: Iterable[str], contains: bool = False
) -> Optional[str]:
    """
    Retorna a primeira coluna cuja normalizacao bate com as chaves.

    :param columns: Colunas disponiveis.
    :param keys: Chaves esperadas.
    :param contains: Se True, permite correspondencia parcial.
    :return: Nome da coluna ou None.
    """
    key_set = set(keys)
    for col in columns:
        norm = _normalize_key(col)
        if norm in key_set:
            return col
        if contains and any(k in norm for k in key_set):
            return col
    return None


def _parse_quarter_from_path(path: Path) -> Optional[Tuple[int, int]]:
    """
    Extrai (ano, trimestre) de um segmento como 3T2025.

    :param path: Caminho do arquivo.
    :return: Tupla (ano, trimestre) ou None.
    """
    match = None
    for part in path.parts:
        match = QUARTER_RE.search(part) or match
    if not match:
        return None
    trimestre = int(match.group(1))
    ano = int(match.group(2))
    return ano, trimestre


def _collect_data_files(base_dir: Path) -> List[Tuple[Path, int, int]]:
    """
    Coleta arquivos elegiveis com ano e trimestre inferidos.

    :param base_dir: Diretorio base de busca.
    :return: Lista de tuplas (arquivo, ano, trimestre).
    """
    files: List[Tuple[Path, int, int]] = []
    for path in base_dir.rglob("*"):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix not in CSV_EXTS and suffix != XLSX_EXT:
            continue
        quarter = _parse_quarter_from_path(path)
        if not quarter:
            continue
        ano, tri = quarter
        files.append((path, ano, tri))
    return files


def _latest_quarters(
    files: List[Tuple[Path, int, int]], limit: int = 3
) -> List[Tuple[Path, int, int]]:
    """
    Filtra apenas os N trimestres mais recentes.

    :param files: Lista de arquivos com ano e trimestre.
    :param limit: Quantidade de trimestres.
    :return: Lista filtrada.
    """
    quarters = sorted({(ano, tri) for _, ano, tri in files}, reverse=True)
    selected = set(quarters[:limit])
    return [item for item in files if (item[1], item[2]) in selected]


def _parse_decimal_series(series: "pd.Series") -> "pd.Series":
    """
    Converte serie monetaria tratando ponto e virgula.

    :param series: Serie com valores monetarios.
    :return: Serie numerica.
    """
    text = series.astype(str)
    has_dot = text.str.contains(r"\.", regex=True)
    has_comma = text.str.contains(",", regex=False)
    text = text.where(~(has_dot & has_comma), text.str.replace(".", "", regex=False))
    text = text.str.replace(",", ".", regex=False)
    return pd.to_numeric(text, errors="coerce").fillna(0.0)


def _should_keep_col(name: str) -> bool:
    """
    Decide se a coluna e necessaria para consolidacao.

    :param name: Nome da coluna.
    :return: True se a coluna deve ser mantida.
    """
    norm = _normalize_key(name)
    if norm in {"regans", "registroans", "descricao"}:
        return True
    return "vlsaldofinal" in norm


def _map_cols(columns: Iterable[str]) -> Optional[Tuple[str, str, str]]:
    """
    Mapeia colunas de REG_ANS, DESCRICAO e VL_SALDO_FINAL.

    :param columns: Colunas disponiveis.
    :return: Tupla (reg, desc, valor) ou None.
    """
    reg_col = _find_col(columns, {"regans", "registroans"})
    desc_col = _find_col(columns, {"descricao"})
    val_col = _find_col(columns, {"vlsaldofinal"}, contains=True)
    if not reg_col or not desc_col or not val_col:
        return None
    return reg_col, desc_col, val_col


def _accumulate_chunk(
    chunk: "pd.DataFrame",
    reg_col: str,
    desc_col: str,
    val_col: str,
    ano: int,
    tri: int,
    agg: Dict[Tuple[str, int, int], float],
) -> None:
    """
    Acumula linhas que casam no agregado do chunk.

    :param chunk: DataFrame do chunk.
    :param reg_col: Coluna de REG_ANS.
    :param desc_col: Coluna de DESCRICAO.
    :param val_col: Coluna de VL_SALDO_FINAL.
    :param ano: Ano do trimestre.
    :param tri: Numero do trimestre.
    :param agg: Dicionario de agregacao.
    :return: None.
    """
    descricao = chunk[desc_col].astype(str).str.lower()
    mask = descricao.str.contains("despesa", na=False) & descricao.str.contains(
        r"(?:evento|sinistro)", na=False
    )
    if not mask.any():
        return
    filtered = chunk.loc[mask, [reg_col, val_col]]
    reg = filtered[reg_col].astype(str).str.strip()
    valores = _parse_decimal_series(filtered[val_col])
    grouped = valores.groupby(reg).sum()
    for reg_ans, total in grouped.items():
        key = (reg_ans, ano, tri)
        agg[key] = agg.get(key, 0.0) + float(total)


def _accumulate_csv(
    path: Path, ano: int, tri: int, agg: Dict[Tuple[str, int, int], float]
) -> None:
    """
    Le CSV/TXT em chunks e acumula valores no agregado.

    :param path: Caminho do arquivo.
    :param ano: Ano do trimestre.
    :param tri: Numero do trimestre.
    :param agg: Dicionario de agregacao.
    :return: None.
    """
    for encoding in CSV_ENCODINGS:
        for sep in CSV_SEPS:
            try:
                reader = pd.read_csv(
                    path,
                    sep=sep,
                    engine="python",
                    dtype=str,
                    chunksize=50000,
                    encoding=encoding,
                    on_bad_lines="skip",
                    usecols=_should_keep_col,
                )
                col_map = None
                for chunk in reader:
                    if col_map is None:
                        col_map = _map_cols(chunk.columns)
                        if not col_map:
                            break
                    reg_col, desc_col, val_col = col_map
                    _accumulate_chunk(chunk, reg_col, desc_col, val_col, ano, tri, agg)
                if col_map:
                    return
            except Exception:
                continue


def _parse_decimal_value(value: object) -> float:
    """
    Converte um valor monetario unico para float.

    :param value: Valor original.
    :return: Valor convertido.
    """
    if value is None:
        return 0.0
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _accumulate_xlsx(
    path: Path, ano: int, tri: int, agg: Dict[Tuple[str, int, int], float]
) -> None:
    """
    Faz streaming de XLSX e acumula linhas no agregado.

    :param path: Caminho do arquivo.
    :param ano: Ano do trimestre.
    :param tri: Numero do trimestre.
    :param agg: Dicionario de agregacao.
    :return: None.
    """
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception:
        return
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            if not header:
                continue
            reg_idx = desc_idx = val_idx = None
            for idx, name in enumerate(header):
                norm = _normalize_key(name)
                if norm in {"regans", "registroans"}:
                    reg_idx = idx
                elif norm == "descricao":
                    desc_idx = idx
                elif "vlsaldofinal" in norm:
                    val_idx = idx
            if reg_idx is None or desc_idx is None or val_idx is None:
                continue
            for row in rows:
                if not row or desc_idx >= len(row):
                    continue
                desc = row[desc_idx]
                if desc is None:
                    continue
                text = str(desc).lower()
                if "despesa" not in text or (
                    "evento" not in text and "sinistro" not in text
                ):
                    continue
                reg = row[reg_idx] if reg_idx < len(row) else None
                val = row[val_idx] if val_idx < len(row) else None
                if reg is None:
                    continue
                reg_ans = str(reg).strip()
                total = _parse_decimal_value(val)
                key = (reg_ans, ano, tri)
                agg[key] = agg.get(key, 0.0) + total
    finally:
        wb.close()


def _load_cadop(path: Path) -> "pd.DataFrame":
    """
    Carrega CADOP com fallback de encoding e colunas normalizadas.

    :param path: Caminho do CADOP.
    :return: DataFrame do CADOP normalizado.
    """
    df = None
    last_error = None
    for enc in CADOP_ENCODINGS:
        try:
            df = pd.read_csv(path, sep=";", encoding=enc, dtype=str)
            break
        except Exception as exc:
            last_error = exc
            continue
    if df is None:
        raise ValueError(f"Falha ao ler CADOP: {last_error}")
    cols = {_normalize_key(c): c for c in df.columns}
    reg_col = cols.get("registrooperadora")
    cnpj_col = cols.get("cnpj")
    razao_col = cols.get("razaosocial")
    data_col = cols.get("dataregistroans")
    if not reg_col or not cnpj_col or not razao_col:
        raise ValueError("Colunas obrigatorias nao encontradas no CADOP.")
    cadop = df[[reg_col, cnpj_col, razao_col] + ([data_col] if data_col else [])].copy()
    cadop.columns = ["REG_ANS", "CNPJ", "RazaoSocial"] + (
        ["DataRegistroANS"] if data_col else []
    )
    if "DataRegistroANS" in cadop.columns:
        cadop["DataRegistroANS"] = pd.to_datetime(
            cadop["DataRegistroANS"], errors="coerce"
        )
    return cadop


def consolidate(
    extract_dir: Path = EXTRACT_DIR,
    cadop_path: Path = CADOP_PATH,
    output_file: Path = OUTPUT_FILE,
    limit_quarters: int = 3,
) -> Path:
    """
    Consolida os ultimos trimestres em um unico CSV.

    :param extract_dir: Diretorio de extracao.
    :param cadop_path: Caminho do CADOP.
    :param output_file: Caminho do CSV de saida.
    :param limit_quarters: Quantidade de trimestres.
    :return: Caminho do CSV consolidado.
    """
    files = _collect_data_files(extract_dir)
    files = _latest_quarters(files, limit=limit_quarters)

    agg: Dict[Tuple[str, int, int], float] = {}
    for path, ano, tri in files:
        if path.suffix.lower() in CSV_EXTS:
            _accumulate_csv(path, ano, tri, agg)
        elif path.suffix.lower() == XLSX_EXT:
            _accumulate_xlsx(path, ano, tri, agg)

    consolidations = []
    for (reg_ans, ano, tri), total in agg.items():
        consolidations.append(
            {
                "REG_ANS": reg_ans,
                "Ano": int(ano),
                "Trimestre": int(tri),
                "ValorDespesas": total,
            }
        )

    consolidated_df = pd.DataFrame(consolidations)
    if consolidated_df.empty:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        consolidated_df.to_csv(output_file, index=False, encoding="utf-8-sig")
        return output_file

    cadop_df = _load_cadop(cadop_path)
    merged_df = consolidated_df.merge(cadop_df, on="REG_ANS", how="left")

    if "DataRegistroANS" in cadop_df.columns:
        cadop_latest = cadop_df.sort_values("DataRegistroANS")
        latest_razao = (
            cadop_latest.dropna(subset=["CNPJ"]).groupby("CNPJ")["RazaoSocial"].last()
        )
        merged_df["RazaoSocial"] = (
            merged_df["CNPJ"].map(latest_razao).fillna(merged_df["RazaoSocial"])
        )

    final_df = merged_df[["CNPJ", "RazaoSocial", "Trimestre", "Ano", "ValorDespesas"]]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    final_df.to_csv(output_file, index=False, encoding="utf-8-sig")
    return output_file


if __name__ == "__main__":
    output = consolidate()
    print(f"CSV gerado: {output}")
